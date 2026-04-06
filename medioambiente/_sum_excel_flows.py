from openpyxl import load_workbook
from pathlib import Path
p=Path(__file__).parent / 'PRÁCTICA MEDIOAMBIENTE(1).xlsx'
wb_vals = load_workbook(p, data_only=True)
wb_form = load_workbook(p, data_only=False)
ws_vals = wb_vals['PRACTICA 1']
ws_form = wb_form['PRACTICA 1']
flows = {}
for row in ws_vals.iter_rows(values_only=False):
    for cell in row:
        v = cell.value
        if isinstance(v, str) and 'Flujo' in v:
            r = cell.row; c = cell.column
            val_right = ws_vals.cell(row=r, column=c+1).value
            formula_right = ws_form.cell(row=r, column=c+1).value
            flows[v.strip()] = (val_right, formula_right, cell.coordinate, ws_vals.cell(row=r, column=c+1).coordinate)
# Also look for labeled intercambios
for row in ws_vals.iter_rows(values_only=False):
    for cell in row:
        v = cell.value
        if isinstance(v, str) and ('intercambio' in v.lower() or 'intercambio' in v.lower()):
            r=cell.row; c=cell.column
            val_right = ws_vals.cell(row=r, column=c+1).value
            formula_right = ws_form.cell(row=r, column=c+1).value
            flows[v.strip()] = (val_right, formula_right, cell.coordinate, ws_vals.cell(row=r, column=c+1).coordinate)
# Print found flows
print('Found flow labels and values:')
for k,(val,formula,coord_val,coord_cell) in flows.items():
    print(f"{k}: value={val!s}, formula={formula!s}, label_cell={coord_val}, value_cell={coord_cell}")
# Also try to find specific expected labels by approximate matching
keys_to_check = ['Flujo Calor Pies','Flujo Calor Cuerpo','Flujo Conveccion','Flujo Cabeza Conveccion','Flujo Radiacion']
print('\nExplicit checks:')
for key in keys_to_check:
    found=False
    for row in ws_vals.iter_rows(values_only=False):
        for cell in row:
            if isinstance(cell.value,str) and cell.value.strip()==key:
                r=cell.row; c=cell.column
                val = ws_vals.cell(row=r, column=c+1).value
                form = ws_form.cell(row=r, column=c+1).value
                print(f"{key}: {val!s}, formula={form!s}, at {ws_vals.cell(row=r,column=c+1).coordinate}")
                found=True
    if not found:
        print(f"{key}: NOT FOUND")
# Search for any numeric cell close to 191 or 200
import math
candidates=[]
for row in ws_vals.iter_rows(values_only=True):
    for v in row:
        try:
            if isinstance(v,(int,float)) and (abs(v-191)<1 or abs(v-200)<1):
                candidates.append(v)
        except Exception:
            pass
print('\nCandidates near 191 or 200:', candidates)
# Compute sum of the main components using values we found
components = []
for name in ['Flujo Calor Pies','Flujo Calor Cuerpo','Flujo Conveccion','Flujo Cabeza Conveccion','Flujo Radiacion']:
    # find value
    val=None
    for row in ws_vals.iter_rows(values_only=False):
        for cell in row:
            if isinstance(cell.value,str) and cell.value.strip()==name:
                val = ws_vals.cell(row=cell.row, column=cell.column+1).value
    components.append((name,val))
print('\nComponent values:')
for n,v in components:
    print(n, v)
sum_components = sum([v for (_,v) in components if isinstance(v,(int,float))])
print('\nSum of main components =', sum_components)
# Check evaporative/respiratory terms in sheet: search for 'intercambio' phrases
extra_names = ['intercambio calor evaporacion','intercambio conveccion respiracion','intercambio evaporacion respiracion']
found_extra=[]
for row in ws_vals.iter_rows(values_only=False):
    for cell in row:
        if isinstance(cell.value,str) and 'intercambio' in cell.value.lower():
            val = ws_vals.cell(row=cell.row, column=cell.column+1).value
            found_extra.append((cell.value.strip(), val, cell.coordinate, ws_vals.cell(row=cell.row,column=cell.column+1).coordinate))
print('\nFound extra terms:')
for item in found_extra:
    print(item)
if found_extra:
    total_with_extra = sum_components + sum([v for (_,v,_,_) in found_extra if isinstance(v,(int,float))])
    print('\nSum including extras =', total_with_extra)
