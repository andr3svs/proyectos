from openpyxl import load_workbook
from pathlib import Path
import math
p=Path(__file__).parent / 'PRÁCTICA MEDIOAMBIENTE(1).xlsx'
wb_form = load_workbook(p, data_only=False)
wb_vals = load_workbook(p, data_only=True)
ws_form = wb_form['PRACTICA 1']
ws_vals = wb_vals['PRACTICA 1']
coords = ['B2','B7','E7','B10','H8','B8']
print('Cell  | value (cached) | formula')
for c in coords:
    v = ws_vals[c].value
    f = ws_form[c].value
    print(f"{c:4} | {str(v):15} | {str(f)}")
B2 = ws_vals['B2'].value
B7 = ws_vals['B7'].value
E7 = ws_vals['E7'].value
B10 = ws_vals['B10'].value
H8 = ws_vals['H8'].value
B8 = ws_vals['B8'].value
print('\nNumeric used for compute:')
print('B2 (T cuerpo)=',B2)
print('B7 (T ropa)=',B7)
print('E7 (k ropa)=',E7)
print('B10 (altura)=',B10)
print('H8 (r_exterior)=',H8)
print('B8 (r_interior)=',B8)
# Compute Q as in excel formula
Q = math.pi * ((B2 - B7) * E7 * 2 * B10) / math.log(H8 / B8)
print('\nQ_excel=',Q)
# Compute using the small radii used in code (0.02 m)
Q_code_radii = math.pi * ((B2 - B7) * E7 * 2 * B10) / math.log((0.02+0.0035)/0.02)
print('Q_with_small_radii=',Q_code_radii)
print('ratio excel/code_radii =',Q/Q_code_radii)
